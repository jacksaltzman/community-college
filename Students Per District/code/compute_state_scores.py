#!/usr/bin/env python3
"""
compute_state_scores.py

Reads the v43.6 spreadsheet's "State Detail Data" sheet, extracts 10 dimension
scores for each state, and merges them into the existing states.json.

Dimension scores (0-100 scale):
  civicEngagementScore, senatorInfluenceScore, filingComplexityScore,
  senatorResponsivenessScore, digitalAdoptionScore, eitcOpportunityScore,
  urbanConcentrationScore, taxDensityScore, youngProfConcentrationScore,
  competitiveDistrictDensityScore
"""

import json
import os
import sys

import openpyxl
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
STATES_JSON = os.path.join(PROJECT_DIR, "app", "public", "data", "states.json")
SPREADSHEET = os.path.join(
    os.path.expanduser("~"),
    "Library",
    "Mobile Documents",
    "com~apple~CloudDocs",
    "Accountable",
    "gtm",
    "State Models",
    "accountable_50_state_analysis_v43_6.xlsx",
)

SHEET_NAME = "State Detail Data"
HEADER_ROW = 3  # Row containing detailed column headers
DATA_START_ROW = 4  # First data row

# ---------------------------------------------------------------------------
# Header-text -> JSON field name mapping
# We match on a normalised substring of the spreadsheet header.
# ---------------------------------------------------------------------------
HEADER_TO_FIELD = {
    "Civic\nEngagement\nScore": "civicEngagementScore",
    "Senator\nInfluence\nScore": "senatorInfluenceScore",
    "Filing\nComplexity\nScore": "filingComplexityScore",
    "Senator\nResponsiveness\nScore": "senatorResponsivenessScore",
    "Digital\nAdoption\nScore": "digitalAdoptionScore",
    "EITC\nOpportunity\nScore": "eitcOpportunityScore",
    "Urban\nConcentration\nScore": "urbanConcentrationScore",
    "Tax Density\nScore": "taxDensityScore",
    "Young Prof\nConcentration\nScore": "youngProfConcentrationScore",
    "CDD\nScore": "competitiveDistrictDensityScore",
}

# ---------------------------------------------------------------------------
# State name -> abbreviation
# ---------------------------------------------------------------------------
STATE_ABBREVS = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT",
    "Delaware": "DE", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI",
    "Minnesota": "MN", "Mississippi": "MS", "Missouri": "MO",
    "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND",
    "Ohio": "OH", "Oklahoma": "OK", "Oregon": "OR",
    "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
    "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
}


def normalise(text: str) -> str:
    """Collapse whitespace for header comparison."""
    return " ".join(str(text).split()).strip().lower()


def main():
    # ------------------------------------------------------------------
    # 1. Open the spreadsheet
    # ------------------------------------------------------------------
    if not os.path.exists(SPREADSHEET):
        print(f"ERROR: Spreadsheet not found at\n  {SPREADSHEET}")
        sys.exit(1)

    wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)
    ws = wb[SHEET_NAME]

    # ------------------------------------------------------------------
    # 2. Build column-index map by matching header text (row 3)
    # ------------------------------------------------------------------
    col_map: dict[str, int] = {}  # field_name -> column index (1-based)

    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(row=HEADER_ROW, column=col_idx).value
        if header_val is None:
            continue
        for header_text, field_name in HEADER_TO_FIELD.items():
            if normalise(header_text) == normalise(header_val):
                col_map[field_name] = col_idx
                break

    # Verify we found all 10
    missing = set(HEADER_TO_FIELD.values()) - set(col_map.keys())
    if missing:
        print(f"ERROR: Could not locate columns for: {missing}")
        print("Headers found:")
        for col_idx in range(1, ws.max_column + 1):
            h = ws.cell(row=HEADER_ROW, column=col_idx).value
            if h:
                print(f"  {get_column_letter(col_idx)}: {repr(h)}")
        sys.exit(1)

    print("Column mapping verified:")
    for field, col_idx in sorted(col_map.items(), key=lambda x: x[1]):
        letter = get_column_letter(col_idx)
        print(f"  {letter} -> {field}")

    # ------------------------------------------------------------------
    # 3. Extract scores for each state
    # ------------------------------------------------------------------
    scores: dict[str, dict[str, float]] = {}

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        state_name = ws.cell(row=row_idx, column=1).value
        if state_name not in STATE_ABBREVS:
            continue  # skip non-state rows (e.g. MAX formulas)

        abbrev = STATE_ABBREVS[state_name]
        state_scores: dict[str, float] = {}

        for field_name, col_idx in col_map.items():
            raw = ws.cell(row=row_idx, column=col_idx).value
            if raw is None:
                state_scores[field_name] = 0.0
            else:
                # Round to 1 decimal for clean JSON
                state_scores[field_name] = round(float(raw), 1)

        scores[abbrev] = state_scores

    print(f"\nExtracted scores for {len(scores)} states")

    if len(scores) != 50:
        print(f"WARNING: Expected 50 states, got {len(scores)}")

    # ------------------------------------------------------------------
    # 4. Load existing states.json and merge
    # ------------------------------------------------------------------
    with open(STATES_JSON, "r") as f:
        states_data = json.load(f)

    merged_count = 0
    for abbrev, state_scores in scores.items():
        if abbrev in states_data:
            states_data[abbrev].update(state_scores)
            merged_count += 1
        else:
            print(f"WARNING: {abbrev} not found in states.json")

    # ------------------------------------------------------------------
    # 5. Write back
    # ------------------------------------------------------------------
    with open(STATES_JSON, "w") as f:
        json.dump(states_data, f, indent=2)
        f.write("\n")

    print(f"Merged 10 dimension scores for {merged_count} states")

    # ------------------------------------------------------------------
    # 6. Validation: top 5 states by each score
    # ------------------------------------------------------------------
    FIELDS = list(HEADER_TO_FIELD.values())
    print(f"\n{'='*80}")
    print("VALIDATION: Top 5 states by composite (avg of 10 scores)")
    print(f"{'='*80}")

    # Compute average across all 10 scores
    avg_scores = {}
    for abbrev, s in scores.items():
        avg_scores[abbrev] = round(sum(s.values()) / len(s), 1)

    top5 = sorted(avg_scores.items(), key=lambda x: x[1], reverse=True)[:5]

    header = f"{'State':<8}"
    for f_name in FIELDS:
        short = f_name.replace("Score", "").replace("senator", "sen").replace("concentration", "conc")
        header += f"{short:>10}"
    header += f"{'AVG':>8}"
    print(header)
    print("-" * len(header))

    for abbrev, avg in top5:
        s = scores[abbrev]
        line = f"{abbrev:<8}"
        for f_name in FIELDS:
            line += f"{s[f_name]:>10.1f}"
        line += f"{avg:>8.1f}"
        print(line)


if __name__ == "__main__":
    main()
